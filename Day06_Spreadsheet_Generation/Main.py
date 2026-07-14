#Importing Necessary Libraries
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
import requests

# Network configuration settings
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3"
}

# Sheet styling definitions
header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
center_alignment = Alignment(horizontal="center", vertical="center")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
red_font = Font(color="9C0006", bold=True)
ALERT_RULE = CellIsRule(operator='greaterThan', formula=['100'], fill=red_fill, font=red_font   )

# Setup workspace
wb = Workbook()
ws = wb.active
row1 = ["Date", "Time", "Category", "Details", "Price(In USD)"]
ws.append(row1)

print("Workbook created and headers added successfully.")

with open("system_logs.txt", "r") as file:
    print("Reading the system_logs.txt file...")
    line_count = sum(1 for line in file)
    print("Total number of lines in the file:", line_count)
    file.close()
#Reading the system_logs.txt
with open("system_logs.txt", "r") as file:
    for i in range(line_count+1):
        print("Reading line", i+1, "of the file...")
        logs = file.readline()
        if "USER_EXPENSE" in logs:
            string_list = logs.split(" INFO - ")
            date = string_list[0].split()[0]
            date = date.replace("[", "")
            time = string_list[0].split()[1]
            time = time.replace("]", "")
            info = string_list[1].split(":",1)
            info.append(info[1].split("|"))
            info.append(info[2][2].split())
            category = info[2][0]
            details = info[2][1]
            price = info[3][0]
            currency = info[3][1]
            # print("Date:", date)
            # print("Time:", time)
            # print("Category:", category)
            # print("Details:", details)
            # print("Price:", price)
            # print("Currency:", currency)
            url = f"https://api.frankfurter.dev/v2/rates?base={currency}&quotes=USD&date={date}"
            requst = requests.get(url, headers=headers, timeout=5)
            data = requst.json()
            rate = data[0]["rate"]
            # print("Exchange Rate for", currency, "to USD on", date, "is:", rate)
            converted_price = float(price) * float(rate)
            # print("Converted Price in USD:", converted_price)
            row = [date, time, category, details, converted_price]
            ws.append(row)
        else:
            print("Ignoring the INVALID_FORMAT_LINE")


last_data_row = ws.max_row
sum_row = last_data_row + 1
col_letter = "E"
ws[f"{col_letter}{sum_row}"] = f"=SUM({col_letter}2:{col_letter}{last_data_row})"
ws[f"{col_letter}{sum_row}"].font = Font(bold=True)

for cell in ws[1][:5]:  
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_alignment


for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value is not None:
            max_len = max(max_len, len(str(cell.value)))
    min_width = 12
    ws.column_dimensions[col_letter].width = max(max_len + 3, min_width)

last_row = ws.max_row
target_range = f"E2:E{last_row}"
ws .conditional_formatting.add(target_range, ALERT_RULE)



wb.save("system_logs.xlsx")
